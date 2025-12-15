"""
Foreplay API - Interfaccia Grafica per Estrazione Transcript
Interfaccia web interattiva per scaricare transcript da boards Foreplay
"""

import streamlit as st
import pandas as pd
import json
import re
from datetime import datetime
from foreplay_client import ForeplayAPIClient
import time

# Configurazione pagina
st.set_page_config(
    page_title="Foreplay Transcript Extractor",
    page_icon="🎬",
    layout="wide"
)

# CSS personalizzato
st.markdown("""
<style>
    .main-header {
        font-size: 3rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 5px;
        padding: 1rem;
        margin: 1rem 0;
    }
    .transcript-box {
        background-color: #f8f9fa;
        border-left: 4px solid #007bff;
        padding: 1rem;
        margin: 1rem 0;
        border-radius: 5px;
    }
    .ad-card {
        background: white;
        border: 1px solid #dee2e6;
        border-radius: 10px;
        padding: 1.5rem;
        margin: 1rem 0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
</style>
""", unsafe_allow_html=True)

# API Key - Load from environment variable
import os
from dotenv import load_dotenv

load_dotenv()
API_KEY = os.getenv("FOREPLAY_API_KEY")

if not API_KEY:
    st.error("⚠️ API Key non configurata! Imposta la variabile d'ambiente FOREPLAY_API_KEY")
    st.stop()


def extract_board_id(url: str) -> str:
    """Estrae l'ID della board dall'URL"""
    # Pattern: https://app.foreplay.co/boards/BOARD_ID
    match = re.search(r'/boards/([a-zA-Z0-9_-]+)', url)
    if match:
        return match.group(1)
    # Se è già solo l'ID
    if re.match(r'^[a-zA-Z0-9_-]+$', url):
        return url
    return None


def get_video_ads_with_transcripts(board_id: str, progress_bar=None, status_text=None):
    """Recupera tutti i video ads con transcript dalla board"""
    client = ForeplayAPIClient(API_KEY)
    
    # Recupera tutti gli ads
    if status_text:
        status_text.text("📋 Recupero ads dalla board...")
    
    ads_response = client.get_board_ads(board_id=board_id, limit=200)
    all_ads = ads_response.get('data', [])
    
    # Filtra solo video ads
    video_ads = [ad for ad in all_ads if ad.get('display_format') == 'video']
    
    if status_text:
        status_text.text(f"🎬 Trovati {len(video_ads)} video ads. Recupero dettagli...")
    
    # Recupera dettagli completi per ogni video
    video_ads_complete = []
    total = len(video_ads)
    
    for i, ad in enumerate(video_ads):
        ad_id = ad.get('id')
        
        try:
            # Recupera dettagli completi
            ad_details = client.get_ad_by_id(ad_id)
            
            # Combina dati
            ad_complete = {**ad, **ad_details}
            video_ads_complete.append(ad_complete)
            
            # Aggiorna progress
            if progress_bar:
                progress_bar.progress((i + 1) / total)
            if status_text:
                status_text.text(f"⏳ Processando {i+1}/{total}: {ad.get('name', 'N/A')[:40]}...")
            
            time.sleep(0.1)
            
        except Exception as e:
            st.warning(f"⚠️ Errore recuperando ad {ad_id}: {e}")
    
    return video_ads_complete


def create_csv_dataframe(video_ads):
    """Crea DataFrame per CSV"""
    rows = []
    
    for ad in video_ads:
        # Converti timestamped_transcription in JSON string (gestisci None)
        timestamped_data = ad.get('timestamped_transcription') or []
        timestamped_json = json.dumps(timestamped_data, ensure_ascii=False)
        
        rows.append({
            'ad_id': ad.get('ad_id', ''),
            'name': ad.get('name', ''),
            'brand_id': ad.get('brand_id', ''),
            'description': ad.get('description', '').replace('<br />', '\n').replace('<br>', '\n'),
            'headline': ad.get('headline', ''),
            'full_transcription': ad.get('full_transcription', ''),
            'timestamped_transcription': timestamped_json,  # JSON array completo
            'video_duration_seconds': ad.get('video_duration', 0),
            'display_format': ad.get('display_format', ''),
            'publisher_platform': ', '.join(ad.get('publisher_platform', [])) if isinstance(ad.get('publisher_platform'), list) else ad.get('publisher_platform', ''),
            'live': ad.get('live', False),
            'video_url': ad.get('video', ''),
            'link_url': ad.get('link_url', ''),
        })
    
    return pd.DataFrame(rows)


def create_timestamped_dataframe(video_ads):
    """Crea DataFrame con timestamp dettagliati"""
    rows = []
    
    for ad in video_ads:
        ad_id = ad.get('ad_id', '')
        name = ad.get('name', '')
        timestamped = ad.get('timestamped_transcription') or []
        
        if timestamped:
            for segment in timestamped:
                rows.append({
                    'ad_id': ad_id,
                    'name': name,
                    'start_time': segment.get('startTime', 0),
                    'end_time': segment.get('endTime', 0),
                    'sentence': segment.get('sentence', '').strip()
                })
    
    return pd.DataFrame(rows)


# ==============================================================================
# INTERFACCIA PRINCIPALE
# ==============================================================================

st.markdown('<h1 class="main-header">🎬 Foreplay Transcript Extractor</h1>', unsafe_allow_html=True)

st.markdown("""
<div style="text-align: center; margin-bottom: 2rem;">
    <p style="font-size: 1.2rem; color: #666;">
        Estrai automaticamente i transcript video da qualsiasi board Foreplay
    </p>
</div>
""", unsafe_allow_html=True)

# Sidebar con info
with st.sidebar:
    st.header("ℹ️ Informazioni")
    st.markdown("""
    ### Come usare:
    1. Inserisci l'URL della board
    2. Clicca su "Estrai Transcript"
    3. Visualizza i risultati
    4. Scarica i file CSV/Excel
    
    ### Formato URL:
    ```
    https://app.foreplay.co/boards/BOARD_ID
    ```
    
    ### Le tue boards:
    - #static_ads_brand
    - #billandchris
    - #focus,brain,mind
    - #safari
    - #longevity
    - #funnel_vsl
    - #debt_scaling_variations
    - #video_podcast_style
    - #hydro_sport_electrolytes
    - #concealed
    """)
    
    st.divider()
    
    # Check crediti
    if st.button("💳 Controlla Crediti"):
        with st.spinner("Controllo..."):
            try:
                client = ForeplayAPIClient(API_KEY)
                usage = client.get_usage()
                st.success(f"Crediti: {usage.get('credits_remaining', 'N/A')}")
            except Exception as e:
                st.error(f"Errore: {e}")

# Input principale
st.markdown("### 🔗 Inserisci il Link della Board")

col1, col2 = st.columns([3, 1])

with col1:
    board_url = st.text_input(
        "URL Board Foreplay",
        placeholder="https://app.foreplay.co/boards/0x0exE7AnMj8i9dgp156",
        help="Inserisci l'URL completo della board o solo l'ID"
    )

with col2:
    st.markdown("<br>", unsafe_allow_html=True)
    extract_button = st.button("🚀 Estrai Transcript", type="primary", use_container_width=True)

# Estrazione board ID
if board_url:
    board_id = extract_board_id(board_url)
    if board_id:
        st.success(f"✅ Board ID rilevato: `{board_id}`")
    else:
        st.error("❌ URL non valido. Usa il formato: https://app.foreplay.co/boards/BOARD_ID")

# Processo di estrazione
if extract_button and board_url:
    board_id = extract_board_id(board_url)
    
    if not board_id:
        st.error("❌ Impossibile estrarre l'ID dalla board. Controlla l'URL.")
    else:
        # Container per il processo
        with st.container():
            st.markdown("---")
            st.markdown("### 📊 Processo di Estrazione")
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            try:
                # Recupera video ads
                video_ads = get_video_ads_with_transcripts(board_id, progress_bar, status_text)
                
                if not video_ads:
                    st.warning("⚠️ Nessun video ad trovato in questa board!")
                else:
                    status_text.text(f"✅ Completato! Recuperati {len(video_ads)} video ads")
                    
                    # Salva in session state
                    st.session_state['video_ads'] = video_ads
                    st.session_state['board_id'] = board_id
                    
                    st.success(f"🎉 Trovati {len(video_ads)} video ads con transcript!")
                    
            except Exception as e:
                st.error(f"❌ Errore durante l'estrazione: {e}")
                import traceback
                st.code(traceback.format_exc())

# Visualizza risultati
if 'video_ads' in st.session_state and st.session_state['video_ads']:
    video_ads = st.session_state['video_ads']
    board_id = st.session_state.get('board_id', 'unknown')
    
    st.markdown("---")
    st.markdown("## 📺 Risultati")
    
    # Statistiche
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Video Ads", len(video_ads))
    with col2:
        with_transcript = sum(1 for ad in video_ads if ad.get('full_transcription'))
        st.metric("Con Transcript", with_transcript)
    with col3:
        total_duration = sum(ad.get('video_duration', 0) for ad in video_ads)
        st.metric("Durata Totale", f"{total_duration:.0f}s")
    with col4:
        total_segments = sum(len(ad.get('timestamped_transcription') or []) for ad in video_ads)
        st.metric("Segmenti Totali", total_segments)
    
    # Tabs per diversi formati
    tab1, tab2, tab3 = st.tabs(["📋 Visualizza Transcript", "📥 Esporta CSV", "📊 Esporta Excel"])
    
    with tab1:
        st.markdown("### 🎬 Video Ads con Transcript")
        
        for i, ad in enumerate(video_ads, 1):
            with st.expander(f"🎥 {i}. {ad.get('name', 'Senza nome')}", expanded=(i==1)):
                col1, col2 = st.columns([2, 1])
                
                with col1:
                    st.markdown(f"**Ad ID:** `{ad.get('ad_id')}`")
                    st.markdown(f"**Brand:** {ad.get('brand_name', 'N/A')}")
                    st.markdown(f"**Durata:** {ad.get('video_duration', 0):.1f} secondi")
                    
                    if ad.get('headline'):
                        st.markdown(f"**Headline:** {ad.get('headline')}")
                    
                    if ad.get('video'):
                        st.markdown(f"**Video:** [Link]({ad.get('video')})")
                
                with col2:
                    if ad.get('thumbnail'):
                        st.image(ad.get('thumbnail'), width=200)
                
                # Description
                if ad.get('description'):
                    st.markdown("**Descrizione:**")
                    desc = ad.get('description', '').replace('<br />', '\n').replace('<br>', '\n')
                    st.text_area("", desc, height=100, key=f"desc_{i}", label_visibility="collapsed")
                
                # Full Transcription
                full_trans = ad.get('full_transcription', '')
                if full_trans:
                    st.markdown("**📝 Transcript Completo:**")
                    
                    # Statistiche transcript
                    word_count = len(full_trans.split())
                    char_count = len(full_trans)
                    st.caption(f"📊 Lunghezza: {char_count:,} caratteri | Parole: ~{word_count:,}")
                    
                    # Mostra tutto il transcript in un text_area scrollabile
                    # Calcola altezza in base alla lunghezza del testo (minimo 200, massimo 800)
                    lines_estimate = len(full_trans) / 80  # ~80 caratteri per riga
                    height = min(max(200, int(lines_estimate * 1.5)), 800)
                    
                    st.text_area(
                        "",
                        full_trans,
                        height=height,
                        key=f"full_transcript_{i}",
                        label_visibility="collapsed",
                        help="Scroll per vedere tutto il transcript"
                    )
                    
                    # Bottone per copiare
                    if st.button(f"📋 Copia Transcript", key=f"copy_{i}"):
                        st.code(full_trans, language=None)
                        st.success("✅ Transcript mostrato sopra - puoi selezionarlo e copiarlo!")
                else:
                    st.info("ℹ️ Nessun transcript disponibile per questo video")
                
                # Timestamped Transcription
                timestamped = ad.get('timestamped_transcription') or []
                if timestamped:
                    st.markdown(f"**⏱️ Segmenti Timestampati:** ({len(timestamped)} segmenti)")
                    
                    with st.expander("Mostra segmenti timestampati"):
                        # Crea DataFrame per visualizzazione
                        segments_data = []
                        for seg in timestamped[:50]:  # Mostra primi 50
                            segments_data.append({
                                'Inizio': f"{seg.get('startTime', 0):.2f}s",
                                'Fine': f"{seg.get('endTime', 0):.2f}s",
                                'Testo': seg.get('sentence', '')
                            })
                        
                        if segments_data:
                            df_segments = pd.DataFrame(segments_data)
                            st.dataframe(df_segments, use_container_width=True, hide_index=True)
                            
                            if len(timestamped) > 50:
                                st.caption(f"... e altri {len(timestamped) - 50} segmenti")
    
    with tab2:
        st.markdown("### 📥 Esporta in CSV")
        
        # EXPORT RAPIDO - Solo 3 campi essenziali
        st.markdown("#### ⚡ Export Rapido (Consigliato)")
        st.info("📋 Export essenziale: Solo **ad_id**, **name** e **full_transcription**")
        
        if st.button("⚡ SCARICA CSV RAPIDO (3 campi)", type="primary", use_container_width=True):
            with st.spinner("Generando CSV rapido..."):
                try:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    
                    # DataFrame semplificato - SOLO 3 CAMPI
                    quick_data = []
                    for ad in video_ads:
                        quick_data.append({
                            'ad_id': ad.get('ad_id', ''),
                            'name': ad.get('name', ''),
                            'full_transcription': ad.get('full_transcription', '')
                        })
                    
                    df_quick = pd.DataFrame(quick_data)
                    
                    # Salva localmente
                    quick_filename = f"transcripts_quick_{timestamp}.csv"
                    df_quick.to_csv(quick_filename, index=False, encoding='utf-8-sig')
                    
                    st.success(f"✅ CSV rapido creato: **{quick_filename}**")
                    
                    # Download button
                    csv_data = df_quick.to_csv(index=False, encoding='utf-8-sig')
                    st.download_button(
                        label="⬇️ SCARICA CSV",
                        data=csv_data,
                        file_name=f"transcripts_{board_id}_{timestamp}.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
                    
                    # Preview con evidenziazione
                    st.markdown("**📊 Preview CSV (prime 3 righe):**")
                    st.dataframe(df_quick.head(3), use_container_width=True)
                    
                    # Info
                    col_info1, col_info2, col_info3 = st.columns(3)
                    with col_info1:
                        st.metric("Righe", len(df_quick))
                    with col_info2:
                        st.metric("Colonne", 3)
                    with col_info3:
                        total_chars = sum(len(str(row['full_transcription'])) for _, row in df_quick.iterrows())
                        st.metric("Caratteri Totali", f"{total_chars:,}")
                    
                except Exception as e:
                    st.error(f"Errore: {e}")
        
        st.markdown("---")
        st.markdown("#### 📊 Export Avanzati (Opzionale)")
        
        # Opzioni di esportazione avanzate
        export_option = st.radio(
            "Seleziona formato CSV avanzato:",
            ["CSV Completo (tutti i campi)", "CSV Timestampato (ogni segmento una riga)", "Entrambi"]
        )
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("📄 Genera CSV", type="primary"):
                with st.spinner("Generando CSV..."):
                    try:
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        
                        if export_option in ["CSV Base (info generali + transcript)", "Entrambi"]:
                            # CSV Base
                            df = create_csv_dataframe(video_ads)
                            csv_filename = f"board_{board_id}_transcripts_{timestamp}.csv"
                            df.to_csv(csv_filename, index=False, encoding='utf-8-sig')
                            
                            st.success(f"✅ CSV creato: {csv_filename}")
                            
                            # Download button
                            csv_data = df.to_csv(index=False, encoding='utf-8-sig')
                            st.download_button(
                                label="⬇️ Scarica CSV Base",
                                data=csv_data,
                                file_name=csv_filename,
                                mime="text/csv"
                            )
                            
                            # Preview
                            st.markdown("**Preview CSV Base:**")
                            st.dataframe(df.head(), use_container_width=True)
                        
                        if export_option in ["CSV Timestampato (ogni segmento una riga)", "Entrambi"]:
                            # CSV Timestampato
                            df_ts = create_timestamped_dataframe(video_ads)
                            csv_ts_filename = f"board_{board_id}_timestamped_{timestamp}.csv"
                            df_ts.to_csv(csv_ts_filename, index=False, encoding='utf-8-sig')
                            
                            st.success(f"✅ CSV timestampato creato: {csv_ts_filename}")
                            
                            # Download button
                            csv_ts_data = df_ts.to_csv(index=False, encoding='utf-8-sig')
                            st.download_button(
                                label="⬇️ Scarica CSV Timestampato",
                                data=csv_ts_data,
                                file_name=csv_ts_filename,
                                mime="text/csv"
                            )
                            
                            # Preview
                            st.markdown("**Preview CSV Timestampato:**")
                            st.dataframe(df_ts.head(20), use_container_width=True)
                        
                    except Exception as e:
                        st.error(f"Errore durante la creazione del CSV: {e}")
        
        with col2:
            if st.button("💾 Salva JSON Completo"):
                with st.spinner("Salvando JSON..."):
                    try:
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        json_filename = f"board_{board_id}_complete_{timestamp}.json"
                        
                        with open(json_filename, 'w', encoding='utf-8') as f:
                            json.dump(video_ads, f, indent=2, ensure_ascii=False)
                        
                        st.success(f"✅ JSON salvato: {json_filename}")
                        
                        # Download button
                        json_data = json.dumps(video_ads, indent=2, ensure_ascii=False)
                        st.download_button(
                            label="⬇️ Scarica JSON",
                            data=json_data,
                            file_name=json_filename,
                            mime="application/json"
                        )
                        
                    except Exception as e:
                        st.error(f"Errore durante il salvataggio JSON: {e}")
    
    with tab3:
        st.markdown("### 📊 Esporta in Excel")
        
        if st.button("📗 Genera File Excel", type="primary"):
            with st.spinner("Creando file Excel..."):
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                    
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    excel_filename = f"board_{board_id}_transcripts_{timestamp}.xlsx"
                    
                    wb = Workbook()
                    
                    # Sheet 1: Info Generali
                    ws1 = wb.active
                    ws1.title = "Transcript Completi"
                    
                    # Header
                    headers = ['ad_id', 'name', 'brand_id', 'description', 'headline', 
                               'full_transcription', 'timestamped_transcription', 'video_duration', 'video_url']
                    ws1.append(headers)
                    
                    # Stile header
                    for cell in ws1[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
                    
                    # Dati
                    for ad in video_ads:
                        # Converti timestamped_transcription in JSON string (gestisci None)
                        timestamped_data = ad.get('timestamped_transcription') or []
                        timestamped_json = json.dumps(timestamped_data, ensure_ascii=False)
                        
                        ws1.append([
                            ad.get('ad_id', ''),
                            ad.get('name', ''),
                            ad.get('brand_id', ''),
                            ad.get('description', '').replace('<br />', '\n'),
                            ad.get('headline', ''),
                            ad.get('full_transcription', ''),
                            timestamped_json,  # JSON array completo
                            ad.get('video_duration', 0),
                            ad.get('video', '')
                        ])
                    
                    # Sheet 2: Timestamp Dettagliati
                    ws2 = wb.create_sheet("Timestamp Dettagliati")
                    ws2.append(['ad_id', 'name', 'start_time', 'end_time', 'sentence'])
                    
                    # Stile header
                    for cell in ws2[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
                    
                    # Dati timestampati
                    for ad in video_ads:
                        timestamped = ad.get('timestamped_transcription') or []
                        for seg in timestamped:
                            ws2.append([
                                ad.get('ad_id', ''),
                                ad.get('name', ''),
                                seg.get('startTime', 0),
                                seg.get('endTime', 0),
                                seg.get('sentence', '')
                            ])
                    
                    # Salva
                    wb.save(excel_filename)
                    
                    st.success(f"✅ Excel creato: {excel_filename}")
                    
                    # Leggi file per download
                    with open(excel_filename, 'rb') as f:
                        excel_data = f.read()
                    
                    st.download_button(
                        label="⬇️ Scarica Excel",
                        data=excel_data,
                        file_name=excel_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    st.info(f"📊 File contiene 2 sheets:\n- Sheet 1: Transcript completi\n- Sheet 2: Timestamp dettagliati")
                    
                except Exception as e:
                    st.error(f"Errore durante la creazione dell'Excel: {e}")
                    import traceback
                    st.code(traceback.format_exc())

# Footer
st.markdown("---")
st.markdown("""
<div style="text-align: center; color: #666; padding: 2rem;">
    <p>🎯 Foreplay API Integration - Made with ❤️ by Wasa</p>
    <p>API Key configurata ✅ | Crediti disponibili: Controlla nella sidebar</p>
</div>
""", unsafe_allow_html=True)

